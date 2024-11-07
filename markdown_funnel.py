import os
import sys
import argparse
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import xlrd
import re
from striprtf.striprtf import rtf_to_text
import mammoth
from dataclasses import dataclass
from typing import List, Dict, Optional
import extract_msg
from email.utils import parseaddr
from pathlib import Path
from tqdm import tqdm

def is_program_generated_html(soup):
    if soup.find_all('div', class_='page') and soup.find('div', class_='page-number'):
        return True
    return False

def convert_title(element):
    return f"# {element.get_text().strip()}\n\n"

def convert_section_header(element):
    return f"## {element.get_text().strip()}\n\n"

def convert_paragraph(element):
    return f"{element.get_text().strip()}\n\n"

def convert_list_item(element):
    return f"- {element.get_text().strip()}\n"

def convert_caption(element):
    return f"*{element.get_text().strip()}*\n\n"

def convert_footnote(element):
    return f"[Footnote: {element.get_text().strip()}]\n\n"

def convert_formula(element):
    return f"${element.get_text().strip()}$\n\n"

def convert_footer(element):
    return f"---\n*{element.get_text().strip()}*\n\n"

def convert_header(element):
    return f"*{element.get_text().strip()}*\n---\n\n"

def convert_page_number(element):
    return f"<!-- {element.get_text().strip()} -->\n\n"

def convert_table(table):
    markdown_table = ""
    rows = table.find_all('tr')
    if not rows:
        return ""

    max_cells = max(len(row.find_all('td')) for row in rows)
    markdown_table += "| " + " | ".join([""] * max_cells) + " |\n"
    markdown_table += "| " + " | ".join(["---"] * max_cells) + " |\n"

    for row in rows:
        cells = row.find_all('td')
        markdown_row = "| " + " | ".join(cell.get_text().strip() for cell in cells)
        markdown_row += " | " * (max_cells - len(cells)) + " |\n"
        markdown_table += markdown_row

    return markdown_table + "\n"

def html_to_markdown(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    markdown_content = ""

    if not is_program_generated_html(soup):
        return None

    for page in soup.find_all('div', class_='page'):
        for element in page.children:
            if element.name == 'div' and 'page-number' in element.get('class', []):
                markdown_content += convert_page_number(element)
            elif element.name == 'h1' and 'title' in element.get('class', []):
                markdown_content += convert_title(element)
            elif element.name == 'h2' and 'section-header' in element.get('class', []):
                markdown_content += convert_section_header(element)
            elif element.name == 'p' and 'text' in element.get('class', []):
                markdown_content += convert_paragraph(element)
            elif element.name == 'li' and 'list-item' in element.get('class', []):
                markdown_content += convert_list_item(element)
            elif element.name == 'figcaption' and 'caption' in element.get('class', []):
                markdown_content += convert_caption(element)
            elif element.name == 'div' and 'footnote' in element.get('class', []):
                markdown_content += convert_footnote(element)
            elif element.name == 'div' and 'formula' in element.get('class', []):
                markdown_content += convert_formula(element)
            elif element.name == 'footer' and 'page-footer' in element.get('class', []):
                markdown_content += convert_footer(element)
            elif element.name == 'header' and 'page-header' in element.get('class', []):
                markdown_content += convert_header(element)
            elif element.name == 'table' and 'data-table' in element.get('class', []):
                markdown_content += convert_table(element)

        if page != soup.find_all('div', class_='page')[-1]:
            markdown_content += "\n---\n\n"

    return markdown_content

def convert_rtf_to_markdown(rtf_path):
    try:
        with open(rtf_path, 'r', encoding='utf-8', errors='ignore') as file:
            rtf_text = file.read()
        
        plain_text = rtf_to_text(rtf_text)
        lines = plain_text.split('\n')
        markdown_lines = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line:
                if i + 1 >= len(lines) or not lines[i + 1].strip():
                    if len(line) < 50:
                        if line.isupper():
                            markdown_lines.append(f"# {line.title()}\n")
                        else:
                            markdown_lines.append(f"## {line}\n")
                    else:
                        markdown_lines.append(line + "\n")
                else:
                    markdown_lines.append(line + "\n")
            else:
                markdown_lines.append("\n")
            i += 1
        
        markdown_content = ""
        in_list = False
        for line in markdown_lines:
            stripped_line = line.strip()
            if stripped_line.startswith(('•', '-', '*', '○', '·', '>', '▪')):
                if not in_list:
                    markdown_content += "\n"
                    in_list = True
                markdown_content += f"- {stripped_line[1:].strip()}\n"
            elif stripped_line.startswith(tuple(f"{i}." for i in range(10))):
                if not in_list:
                    markdown_content += "\n"
                    in_list = True
                number, text = stripped_line.split('.', 1)
                markdown_content += f"{number}. {text.strip()}\n"
            else:
                if in_list and stripped_line:
                    markdown_content += "\n"
                in_list = False
                markdown_content += line
        
        return markdown_content
        
    except Exception as e:
        print(f"Error converting RTF file {rtf_path}: {str(e)}")
        return None

def convert_docx_to_markdown(docx_path):
    try:
        style_map = """
        p[style-name='heading 1'] => h1:fresh
        p[style-name='heading 2'] => h2:fresh
        p[style-name='heading 3'] => h3:fresh
        p[style-name='heading 4'] => h4:fresh
        p[style-name='heading 5'] => h5:fresh
        p[style-name='heading 6'] => h6:fresh
        p[style-name='Title'] => h1:fresh
        p[style-name='Subtitle'] => h2:fresh
        p[style-name='List Bullet'] => ul > li:fresh
        p[style-name='List Number'] => ol > li:fresh
        p[style-name='Quote'] => blockquote:fresh
        p[style-name='Intense Quote'] => blockquote.intense:fresh
        p[style-name='Normal'] => p:fresh
        b => strong
        i => em
        strike => s
        u => u
        """
        
        with open(docx_path, "rb") as docx_file:
            result = mammoth.convert_to_markdown(
                docx_file,
                style_map=style_map,
                ignore_empty_paragraphs=True
            )
            
            important_messages = [
                msg for msg in result.messages 
                if "style" not in msg.message.lower() 
                and "message" in msg.message.lower()
            ]
            
            if important_messages:
                print(f"Important messages for {docx_path}:")
                for message in important_messages:
                    print(f"- {message}")
            
            markdown_content = result.value
            markdown_content = re.sub(r'\n{3,}', '\n\n', markdown_content)
            markdown_content = re.sub(r'(\n#{1,6}\s.*)\n(?=\S)', r'\1\n\n', markdown_content)
            markdown_content = re.sub(r'(\n[-*]\s.*)\n(?=[-*]\s)', r'\1\n', markdown_content)
            
            return markdown_content.strip() + '\n'
            
    except Exception as e:
        print(f"Error converting DOCX file {docx_path}: {str(e)}")
        return None

def convert_msg_to_markdown(msg_path):
    try:
        msg = extract_msg.Message(msg_path)
        markdown_parts = []
        
        markdown_parts.append("# " + (msg.subject or "No Subject"))
        markdown_parts.append("\n## Email Details\n")
        
        markdown_parts.append("| Field | Value |")
        markdown_parts.append("|-------|-------|")
        
        sender_name, sender_email = parseaddr(msg.sender)
        from_value = f"{sender_name} <{sender_email}>" if sender_name else sender_email
        markdown_parts.append(f"| From | {from_value} |")
        
        if msg.to:
            to_addresses = msg.to
            if isinstance(to_addresses, str):
                to_addresses = [to_addresses]
            markdown_parts.append(f"| To | {'; '.join(to_addresses)} |")
        
        if msg.cc:
            cc_addresses = msg.cc
            if isinstance(cc_addresses, str):
                cc_addresses = [cc_addresses]
            markdown_parts.append(f"| CC | {'; '.join(cc_addresses)} |")
        
        if msg.date:
            try:
                date_str = msg.date.strftime("%Y-%m-%d %H:%M:%S")
                markdown_parts.append(f"| Date | {date_str} |")
            except:
                pass
        
        markdown_parts.append("")
        
        if msg.body:
            markdown_parts.append("\n## Message Body\n")
            body_text = msg.body
            body_lines = body_text.split('\n')
            processed_lines = []
            
            for line in body_lines:
                if line.strip().startswith('>'):
                    processed_lines.append(line)
                elif line.strip().startswith('----'):
                    processed_lines.append('\n---\n')
                else:
                    processed_lines.append(line)
            
            body_markdown = '\n'.join(processed_lines)
            body_markdown = re.sub(r'\n{3,}', '\n\n', body_markdown)
            markdown_parts.append(body_markdown)
        
        if msg.attachments:
            markdown_parts.append("\n## Attachments\n")
            for attachment in msg.attachments:
                filename = attachment.longFilename or attachment.shortFilename
                if filename:
                    markdown_parts.append(f"- 📎 `{filename}`")
            markdown_parts.append("")
        
        msg.close()
        
        markdown_content = '\n'.join(markdown_parts)
        markdown_content = markdown_content.replace('\r', '')
        markdown_content = re.sub(r'\n{3,}', '\n\n', markdown_content)
        
        return markdown_content.strip() + '\n'
        
    except Exception as e:
        print(f"Error converting MSG file {msg_path}: {str(e)}")
        return None

def convert_excel_to_markdown(excel_path):
    try:
        file_extension = os.path.splitext(excel_path)[1].lower()
        markdown_content = []
        
        if file_extension == '.xls':
            workbook = xlrd.open_workbook(excel_path, formatting_info=True)
            
            for sheet_name in workbook.sheet_names():
                sheet_content = []
                
                if len(workbook.sheet_names()) > 1:
                    sheet_content.append(f"## Sheet: {sheet_name}\n")
                
                worksheet = workbook.sheet_by_name(sheet_name)
                
                if worksheet.nrows == 0 or worksheet.ncols == 0:
                    sheet_content.append("*Empty sheet*\n")
                    markdown_content.extend(sheet_content)
                    continue
                
                try:
                    data = []
                    for row_idx in range(worksheet.nrows):
                        row_data = []
                        for col_idx in range(worksheet.ncols):
                            cell = worksheet.cell(row_idx, col_idx)
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    value = xlrd.xldate_as_datetime(cell.value, workbook.datemode).strftime('%Y-%m-%d')
                                except Exception:
                                    value = str(cell.value)
                            else:
                                value = str(cell.value).strip()
                            row_data.append(value)
                        data.append(row_data)
                    
                    if data:
                        if len(data[0]) > 0:
                            df = pd.DataFrame(data[1:], columns=data[0])
                            table = df.to_markdown(index=False)
                            sheet_content.append(table + "\n")
                        else:
                            sheet_content.append("*Empty table structure*\n")
                    else:
                        sheet_content.append("*No data found*\n")
                    
                except Exception as e:
                    print(f"Error converting sheet {sheet_name}: {str(e)}")
                    sheet_content.append(f"*Error converting sheet: {str(e)}*\n")
                
                markdown_content.extend(sheet_content)
                
                if sheet_name != workbook.sheet_names()[-1]:
                    markdown_content.append("---\n")
            
        else:  # .xlsx files
            xlsx = pd.ExcelFile(excel_path)
            
            for sheet_name in xlsx.sheet_names:
                sheet_content = []
                
                if len(xlsx.sheet_names) > 1:
                    sheet_content.append(f"## Sheet: {sheet_name}\n")
                
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                
                if df.empty:
                    sheet_content.append("*Empty sheet*\n")
                    markdown_content.extend(sheet_content)
                    continue
                
                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb[sheet_name]
                    
                    merged_cells = {}
                    
                    for merged_range in ws.merged_cells.ranges:
                        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        content = str(top_left_cell.value) if top_left_cell.value is not None else ""
                        
                        for row in range(merged_range.min_row, merged_range.max_row + 1):
                            for col in range(merged_range.min_col, merged_range.max_col + 1):
                                if row == merged_range.min_row and col == merged_range.min_col:
                                    merged_cells[(row, col)] = content
                                else:
                                    merged_cells[(row, col)] = ""
                    
                    new_data = []
                    for row_idx in range(1, ws.max_row + 1):
                        row_data = []
                        for col_idx in range(1, ws.max_column + 1):
                            if (row_idx, col_idx) in merged_cells:
                                cell_content = merged_cells[(row_idx, col_idx)]
                            else:
                                cell = ws.cell(row_idx, col_idx)
                                cell_content = str(cell.value) if cell.value is not None else ""
                            row_data.append(cell_content)
                        new_data.append(row_data)
                    
                    if new_data:
                        if len(new_data[0]) > 0:
                            new_df = pd.DataFrame(new_data[1:], columns=new_data[0])
                            table = new_df.to_markdown(index=False)
                            sheet_content.append(table + "\n")
                        else:
                            sheet_content.append("*Empty table structure*\n")
                    else:
                        sheet_content.append("*No data found*\n")
                    
                    wb.close()
                    
                except Exception as e:
                    print(f"Error converting sheet {sheet_name}: {str(e)}")
                    sheet_content.append(f"*Error converting sheet: {str(e)}*\n")
                
                markdown_content.extend(sheet_content)
                
                if sheet_name != xlsx.sheet_names[-1]:
                    markdown_content.append("---\n")
        
        return "\n".join(markdown_content)
        
    except Exception as e:
        print(f"Error converting Excel file {excel_path}: {str(e)}")
        return None

def process_directory(input_dir):
    """Process all supported files in a directory."""
    # Get list of all files to process
    files_to_process = []
    for root, _, files in os.walk(input_dir):
        for file in files:
            file_path = os.path.join(root, file)
            # Skip files that already have markdown versions
            md_path = os.path.splitext(file_path)[0] + '.md'
            if os.path.exists(md_path):
                print(f"Skipping {file_path} as markdown file already exists")
                continue
                
            if file.lower().endswith('.html'):
                files_to_process.append((file_path, 'html'))
            elif file.lower().endswith(('.xlsx', '.xls')):
                files_to_process.append((file_path, 'excel'))
            elif file.lower().endswith('.rtf'):
                files_to_process.append((file_path, 'rtf'))
            elif file.lower().endswith('.docx') and not file.startswith('~$'):
                files_to_process.append((file_path, 'docx'))
            elif file.lower().endswith('.msg'):
                files_to_process.append((file_path, 'msg'))

    if not files_to_process:
        print("No files found to process")
        return

    print(f"Found {len(files_to_process)} files to process")
    
    # Process each file
    successful = 0
    failed = 0
    
    for file_path, file_type in tqdm(files_to_process, desc="Converting files"):
        try:
            md_path = os.path.splitext(file_path)[0] + '.md'
            markdown_content = None
            
            if file_type == 'html':
                pdf_path = os.path.splitext(file_path)[0] + '.pdf'
                if not os.path.exists(pdf_path):
                    print(f"Skipping {file_path} as no corresponding PDF found")
                    failed += 1
                    continue

                with open(file_path, 'r', encoding='utf-8') as html_file:
                    html_content = html_file.read()
                markdown_content = html_to_markdown(html_content)
                
            elif file_type == 'excel':
                markdown_content = convert_excel_to_markdown(file_path)
                
            elif file_type == 'rtf':
                markdown_content = convert_rtf_to_markdown(file_path)
                
            elif file_type == 'docx':
                markdown_content = convert_docx_to_markdown(file_path)
                
            elif file_type == 'msg':
                markdown_content = convert_msg_to_markdown(file_path)

            if markdown_content:
                # Ensure directory exists
                os.makedirs(os.path.dirname(os.path.abspath(md_path)), exist_ok=True)
                
                # Write the markdown file
                with open(md_path, 'w', encoding='utf-8') as md_file:
                    md_file.write(markdown_content)
                
                # print(f"Converted {file_type.upper()}: {file_path} to {md_path}")
                successful += 1
            else:
                print(f"Failed to convert {file_path}")
                failed += 1
                
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
            failed += 1
    
    print(f"\nConversion complete: {successful} successful, {failed} failed")

def main():
    parser = argparse.ArgumentParser(
        description="Convert HTML, Excel, RTF, DOCX, and MSG files to Markdown"
    )
    parser.add_argument("input_dir", help="Path to the input directory containing files to convert")
    args = parser.parse_args()

    if not os.path.isdir(args.input_dir):
        print(f"Error: {args.input_dir} is not a valid directory.")
        sys.exit(1)

    process_directory(args.input_dir)

if __name__ == "__main__":
    main()
